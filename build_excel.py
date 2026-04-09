import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import re

# ── Load data ────────────────────────────────────────────────────────────────
offers_raw = pd.read_csv('/sessions/wizardly-beautiful-fermi/master_offers_clean.csv')
purchases_raw = pd.read_csv('/sessions/wizardly-beautiful-fermi/master_purchases.csv')

# Remove duplicate headers
offers = offers_raw[offers_raw['Date'] != 'Date'].copy()
purchases = purchases_raw[purchases_raw['Order_Date'] != 'Order_Date'].copy()

# Clean types
for col in ['Price_CHF', 'Original_Price_CHF']:
    offers[col] = pd.to_numeric(offers[col], errors='coerce')
offers['Vintage'] = pd.to_numeric(offers['Vintage'], errors='coerce').astype('Int64')
offers['Date'] = pd.to_datetime(offers['Date'], errors='coerce')
offers = offers.dropna(subset=['Date'])
offers = offers.sort_values(['Wine', 'Vintage', 'Date'])

for col in ['Qty_Bottles', 'Price_Per_Bottle_CHF']:
    purchases[col] = pd.to_numeric(purchases[col], errors='coerce')
purchases['Vintage'] = pd.to_numeric(purchases['Vintage'], errors='coerce').astype('Int64')
purchases['Order_Date'] = pd.to_datetime(purchases['Order_Date'], errors='coerce')
purchases['Total_CHF'] = purchases['Qty_Bottles'] * purchases['Price_Per_Bottle_CHF']

# ── Purchase matching ─────────────────────────────────────────────────────────
def normalize(s):
    if pd.isna(s): return ''
    return re.sub(r'[^a-z0-9]', '', str(s).lower())

# Build set of (norm_wine, vintage) tuples from purchases
purchase_keys = set()
for _, row in purchases.iterrows():
    nw = normalize(row['Wine'])
    nv = str(int(row['Vintage'])) if pd.notna(row['Vintage']) else ''
    if nw:
        purchase_keys.add((nw, nv))
        # Also add first 8 chars for partial matching
        purchase_keys.add((nw[:8], nv))

def is_purchased(wine, vintage):
    nw = normalize(wine)
    nv = str(int(vintage)) if pd.notna(vintage) and vintage == vintage else ''
    if (nw, nv) in purchase_keys: return True
    if (nw[:8], nv) in purchase_keys: return True
    # Partial: any purchase key whose wine starts with the offer wine's prefix
    for pk_wine, pk_vint in purchase_keys:
        if pk_vint == nv and len(nw) >= 5 and (nw[:6] in pk_wine or pk_wine[:6] in nw):
            return True
    return False

offers['Purchased'] = offers.apply(lambda r: is_purchased(r['Wine'], r['Vintage']), axis=1)

# ── Price evolution calc ──────────────────────────────────────────────────────
offers = offers.sort_values(['Wine', 'Vintage', 'Date'])
offers['Prev_Price'] = offers.groupby(['Wine', 'Vintage'])['Price_CHF'].shift(1)
offers['Price_Change_CHF'] = offers['Price_CHF'] - offers['Prev_Price']
offers['Price_Change_Pct'] = (offers['Price_Change_CHF'] / offers['Prev_Price'])
offers['Offer_Count'] = offers.groupby(['Wine', 'Vintage']).cumcount() + 1

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill('solid', start_color='1F3864')   # dark navy
HEADER_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SUBHDR_FILL   = PatternFill('solid', start_color='2E75B6')   # medium blue
SUBHDR_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT     = Font(name='Arial', size=9)
ALT_FILL      = PatternFill('solid', start_color='EBF3FA')   # light blue
YES_FILL      = PatternFill('solid', start_color='C6EFCE')   # light green
YES_FONT      = Font(name='Arial', size=9, bold=True, color='276221')
NO_FONT       = Font(name='Arial', size=9, color='888888')
UP_FONT       = Font(name='Arial', size=9, color='C00000')
DOWN_FONT     = Font(name='Arial', size=9, color='375623')
BORDER_THIN   = Border(
    bottom=Side(style='thin', color='BDD7EE'),
    right=Side(style='thin', color='BDD7EE')
)
WRAP_ALIGN    = Alignment(wrap_text=False, vertical='center')

def set_header(cell, text):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_subheader(cell, text):
    cell.value = text
    cell.font = SUBHDR_FONT
    cell.fill = SUBHDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')

def style_body(cell, alt=False, align='left'):
    cell.font = BODY_FONT
    if alt:
        cell.fill = ALT_FILL
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = BORDER_THIN

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════
# SHEET 1: All Offers
# ═══════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = 'All Offers'

headers1 = ['Date', 'Wine', 'Producer', 'Appellation', 'Vintage',
            'Offer Price\n(CHF/bt)', 'List Price\n(CHF/bt)', 'Discount\n(%)',
            'Prev. Offer\n(CHF/bt)', 'Price Δ\n(CHF)', 'Price Δ\n(%)',
            'Times\nOffered', 'Purchased', 'Email Subject']
col_widths1 = [12, 38, 28, 24, 8, 12, 12, 10, 12, 10, 10, 8, 10, 50]

ws1.row_dimensions[1].height = 35
for ci, (h, w) in enumerate(zip(headers1, col_widths1), 1):
    cell = ws1.cell(row=1, column=ci)
    set_header(cell, h)
    ws1.column_dimensions[get_column_letter(ci)].width = w

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f'A1:{get_column_letter(len(headers1))}1'

for ri, (_, row) in enumerate(offers.iterrows(), 2):
    alt = ri % 2 == 0
    date_val = row['Date'].date() if pd.notna(row['Date']) else ''
    vintage_val = int(row['Vintage']) if pd.notna(row['Vintage']) else ''
    price_val = row['Price_CHF'] if pd.notna(row['Price_CHF']) else ''
    list_price = row['Original_Price_CHF'] if pd.notna(row['Original_Price_CHF']) else ''
    disc_pct = ((list_price - price_val) / list_price) if (list_price and price_val) else ''
    prev_price = row['Prev_Price'] if pd.notna(row['Prev_Price']) else ''
    price_chg = row['Price_Change_CHF'] if pd.notna(row['Price_Change_CHF']) else ''
    price_pct = row['Price_Change_Pct'] if pd.notna(row['Price_Change_Pct']) else ''
    offer_count = int(row['Offer_Count']) if pd.notna(row['Offer_Count']) else 1
    purchased = 'YES' if row['Purchased'] else ''

    values = [date_val, str(row['Wine']) if pd.notna(row['Wine']) else '',
              str(row['Producer']) if pd.notna(row['Producer']) else '',
              str(row['Appellation']) if pd.notna(row['Appellation']) else '',
              vintage_val, price_val, list_price, disc_pct,
              prev_price, price_chg, price_pct, offer_count, purchased,
              str(row['Email_Subject'])[:60] if pd.notna(row['Email_Subject']) else '']

    for ci, val in enumerate(values, 1):
        cell = ws1.cell(row=ri, column=ci, value=val)
        style_body(cell, alt)
        # Special formatting
        if ci == 1: cell.number_format = 'YYYY-MM-DD'
        elif ci in (6, 7, 9, 10): cell.number_format = '#,##0.00'
        elif ci in (8, 11): cell.number_format = '0.0%;-0.0%;"-"'
        elif ci == 13:  # Purchased
            if val == 'YES':
                cell.font = YES_FONT
                cell.fill = YES_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = NO_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center')
        elif ci == 10 and isinstance(val, (int, float)) and val != '':  # Price delta
            if val > 0: cell.font = UP_FONT
            elif val < 0: cell.font = DOWN_FONT

print(f'Sheet 1 written: {len(offers)} rows')

# ═══════════════════════════════════════════════════════════════════
# SHEET 2: Price Evolution (wines offered 2+ times)
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Price Evolution')

# Filter wines offered more than once
multi = offers[offers.groupby(['Wine', 'Vintage'])['Wine'].transform('count') > 1].copy()
multi = multi.sort_values(['Wine', 'Vintage', 'Date'])

headers2 = ['Wine', 'Producer', 'Appellation', 'Vintage',
            'Offer Date', 'Price (CHF/bt)', 'List Price (CHF/bt)', 'Discount (%)',
            'Prev. Price (CHF)', 'Change (CHF)', 'Change (%)', 'Purchased']
col_widths2 = [38, 28, 24, 8, 12, 14, 14, 10, 14, 11, 10, 10]

ws2.row_dimensions[1].height = 30
for ci, (h, w) in enumerate(zip(headers2, col_widths2), 1):
    cell = ws2.cell(row=1, column=ci)
    set_header(cell, h)
    ws2.column_dimensions[get_column_letter(ci)].width = w

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers2))}1'

GROUP_FILLS = [
    PatternFill('solid', start_color='DDEEFF'),
    PatternFill('solid', start_color='F0F7FF'),
]

prev_key = None
group_idx = 0
for ri, (_, row) in enumerate(multi.iterrows(), 2):
    curr_key = (str(row['Wine']), str(row['Vintage']))
    if curr_key != prev_key:
        group_idx = (group_idx + 1) % 2
        prev_key = curr_key
    fill = GROUP_FILLS[group_idx]

    date_val = row['Date'].date() if pd.notna(row['Date']) else ''
    vintage_val = int(row['Vintage']) if pd.notna(row['Vintage']) else ''
    price_val = row['Price_CHF'] if pd.notna(row['Price_CHF']) else ''
    list_price = row['Original_Price_CHF'] if pd.notna(row['Original_Price_CHF']) else ''
    disc_pct = ((list_price - price_val) / list_price) if (list_price and price_val) else ''
    prev_price = row['Prev_Price'] if pd.notna(row['Prev_Price']) else ''
    price_chg = row['Price_Change_CHF'] if pd.notna(row['Price_Change_CHF']) else ''
    price_pct = row['Price_Change_Pct'] if pd.notna(row['Price_Change_Pct']) else ''
    purchased = 'YES' if row['Purchased'] else ''

    values = [str(row['Wine']) if pd.notna(row['Wine']) else '',
              str(row['Producer']) if pd.notna(row['Producer']) else '',
              str(row['Appellation']) if pd.notna(row['Appellation']) else '',
              vintage_val, date_val, price_val, list_price, disc_pct,
              prev_price, price_chg, price_pct, purchased]

    for ci, val in enumerate(values, 1):
        cell = ws2.cell(row=ri, column=ci, value=val)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = BORDER_THIN
        if ci == 5: cell.number_format = 'YYYY-MM-DD'
        elif ci in (6, 7, 9, 10): cell.number_format = '#,##0.00'
        elif ci in (8, 11): cell.number_format = '0.0%;-0.0%;"-"'
        elif ci == 12:
            if val == 'YES':
                cell.font = YES_FONT
                cell.fill = YES_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
        elif ci == 10 and isinstance(val, (int, float)) and val != '':
            if val > 0: cell.font = Font(name='Arial', size=9, color='C00000')
            elif val < 0: cell.font = Font(name='Arial', size=9, color='375623')

print(f'Sheet 2 written: {len(multi)} rows ({multi.groupby(["Wine","Vintage"]).ngroups} unique wine/vintages)')

# ═══════════════════════════════════════════════════════════════════
# SHEET 3: My Purchases
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('My Purchases')

pur = purchases.copy()
pur['Total_CHF'] = pur['Qty_Bottles'] * pur['Price_Per_Bottle_CHF']
pur = pur.sort_values('Order_Date', ascending=False)

headers3 = ['Order Date', 'Wine', 'Producer', 'Appellation', 'Vintage',
            'Qty\n(Bottles)', 'Price/bt\n(CHF)', 'Total\n(CHF)', 'Order Note']
col_widths3 = [12, 38, 28, 24, 8, 8, 10, 10, 55]

ws3.row_dimensions[1].height = 35
for ci, (h, w) in enumerate(zip(headers3, col_widths3), 1):
    cell = ws3.cell(row=1, column=ci)
    set_header(cell, h)
    ws3.column_dimensions[get_column_letter(ci)].width = w

ws3.freeze_panes = 'A2'
ws3.auto_filter.ref = f'A1:{get_column_letter(len(headers3))}1'

for ri, (_, row) in enumerate(pur.iterrows(), 2):
    alt = ri % 2 == 0
    date_val = row['Order_Date'].date() if pd.notna(row['Order_Date']) else ''
    vintage_val = int(row['Vintage']) if pd.notna(row['Vintage']) else ''
    qty = row['Qty_Bottles'] if pd.notna(row['Qty_Bottles']) else ''
    price = row['Price_Per_Bottle_CHF'] if pd.notna(row['Price_Per_Bottle_CHF']) else ''
    total = row['Total_CHF'] if pd.notna(row['Total_CHF']) else ''

    values = [date_val,
              str(row['Wine']) if pd.notna(row['Wine']) else '',
              str(row['Producer']) if pd.notna(row['Producer']) else '',
              str(row['Appellation']) if pd.notna(row['Appellation']) else '',
              vintage_val, qty, price, total,
              str(row['Order_Note'])[:80] if pd.notna(row['Order_Note']) else '']

    for ci, val in enumerate(values, 1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        style_body(cell, alt, 'left')
        if ci == 1: cell.number_format = 'YYYY-MM-DD'
        elif ci in (7, 8): cell.number_format = '#,##0.00'
        elif ci == 6: cell.alignment = Alignment(horizontal='center', vertical='center')

# Total row
total_row = len(pur) + 2
ws3.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Arial', bold=True, size=10)
ws3.cell(row=total_row, column=8).value = f'=SUM(H2:H{total_row-1})'
ws3.cell(row=total_row, column=8).font = Font(name='Arial', bold=True, size=10)
ws3.cell(row=total_row, column=8).number_format = '#,##0.00'
ws3.cell(row=total_row, column=8).fill = PatternFill('solid', start_color='1F3864')
ws3.cell(row=total_row, column=8).font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
ws3.cell(row=total_row, column=6).value = f'=SUM(F2:F{total_row-1})'
ws3.cell(row=total_row, column=6).font = Font(name='Arial', bold=True, size=10)
ws3.cell(row=total_row, column=6).number_format = '#,##0'
ws3.cell(row=total_row, column=6).fill = PatternFill('solid', start_color='1F3864')
ws3.cell(row=total_row, column=6).font = Font(name='Arial', bold=True, size=10, color='FFFFFF')

print(f'Sheet 3 written: {len(pur)} purchase rows')

# ═══════════════════════════════════════════════════════════════════
# SHEET 4: Summary Dashboard
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Summary', 0)  # First sheet

ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 18
ws4.column_dimensions['C'].width = 30

# Title
ws4.merge_cells('A1:C1')
ws4['A1'] = 'EDULIS WINE DATABASE'
ws4['A1'].font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
ws4['A1'].fill = PatternFill('solid', start_color='1F3864')
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 40

ws4.merge_cells('A2:C2')
ws4['A2'] = f'Generated from {len(offers):,} offer emails  ·  Source: edulis.ch'
ws4['A2'].font = Font(name='Arial', italic=True, size=10, color='2E75B6')
ws4['A2'].alignment = Alignment(horizontal='center')
ws4.row_dimensions[2].height = 20

# Section: Offers
def dash_header(ws, row, text, cols='A:C'):
    ws.merge_cells(f'{cols.split(":")[0]}{row}:{cols.split(":")[1]}{row}')
    cell = ws[f'{cols.split(":")[0]}{row}']
    cell.value = text
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = SUBHDR_FILL
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22

def dash_row(ws, row, label, value, fmt='', note=''):
    ws[f'A{row}'] = label
    ws[f'A{row}'].font = Font(name='Arial', size=10)
    ws[f'A{row}'].alignment = Alignment(indent=2, vertical='center')
    ws[f'B{row}'] = value
    ws[f'B{row}'].font = Font(name='Arial', bold=True, size=10, color='1F3864')
    ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
    if fmt: ws[f'B{row}'].number_format = fmt
    if note:
        ws[f'C{row}'] = note
        ws[f'C{row}'].font = Font(name='Arial', size=9, color='888888', italic=True)
    ws.row_dimensions[row].height = 18
    # Alternating bg
    if row % 2 == 0:
        for col in 'ABC':
            ws[f'{col}{row}'].fill = PatternFill('solid', start_color='EBF3FA')

r = 4
dash_header(ws4, r, '  OFFER OVERVIEW')
r += 1
dash_row(ws4, r, 'Total offer emails processed', len(offers), '#,##0')
r += 1
dash_row(ws4, r, 'Unique wines & vintages', int(offers.dropna(subset=['Wine','Vintage']).groupby(['Wine','Vintage']).ngroups), '#,##0')
r += 1
dash_row(ws4, r, 'Wines offered multiple times', int(multi.groupby(['Wine','Vintage']).ngroups), '#,##0')
r += 1
dash_row(ws4, r, 'Date range (oldest offer)', str(offers['Date'].min().date()), '', 'Oldest in extracted dataset')
r += 1
dash_row(ws4, r, 'Date range (most recent)', str(offers['Date'].max().date()))
r += 1
dash_row(ws4, r, 'Offers with discount (list price shown)', int(offers['Original_Price_CHF'].notna().sum()), '#,##0')
r += 1
dash_row(ws4, r, 'Average offer price (CHF/bt)', round(float(offers['Price_CHF'].mean()), 1), '#,##0.0')
r += 1
dash_row(ws4, r, 'Min offer price (CHF/bt)', float(offers['Price_CHF'].min()), '#,##0.0')
r += 1
dash_row(ws4, r, 'Max offer price (CHF/bt)', float(offers['Price_CHF'].max()), '#,##0.0')

r += 2
dash_header(ws4, r, '  PURCHASE HISTORY')
r += 1
total_bottles = purchases['Qty_Bottles'].sum()
total_spend = (purchases['Qty_Bottles'] * purchases['Price_Per_Bottle_CHF']).sum()
dash_row(ws4, r, 'Total orders placed', len(purchases), '#,##0')
r += 1
dash_row(ws4, r, 'Total bottles purchased', int(total_bottles) if pd.notna(total_bottles) else 'n/a', '#,##0')
r += 1
dash_row(ws4, r, 'Total estimated spend (CHF)', round(float(total_spend), 0) if pd.notna(total_spend) else 'n/a', '#,##0', 'excl. TVA, based on order prices')
r += 1
dash_row(ws4, r, 'Earliest purchase on record', str(purchases['Order_Date'].min().date()) if pd.notna(purchases['Order_Date'].min()) else 'n/a')
r += 1
dash_row(ws4, r, 'Most recent purchase', str(purchases['Order_Date'].max().date()) if pd.notna(purchases['Order_Date'].max()) else 'n/a')

# Top producers in purchases
r += 2
dash_header(ws4, r, '  TOP PRODUCERS (by orders)')
r += 1
top_prod = purchases['Producer'].value_counts().head(10)
for prod, cnt in top_prod.items():
    dash_row(ws4, r, f'  {prod}', cnt, '#,##0', 'orders')
    r += 1

# Top wines by number of offers
r += 1
dash_header(ws4, r, '  MOST FREQUENTLY OFFERED WINES')
r += 1
top_wines = offers.groupby(['Wine', 'Vintage'])['Wine'].count().sort_values(ascending=False).head(10)
for (wine, vint), cnt in top_wines.items():
    vstr = str(int(vint)) if pd.notna(vint) else 'NV'
    dash_row(ws4, r, f'  {wine} {vstr}', cnt, '#,##0', 'times offered')
    r += 1

ws4.freeze_panes = 'A3'

print('Sheet 4 (Summary) written')

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = '/sessions/wizardly-beautiful-fermi/mnt/outputs/Edulis_Wine_Database.xlsx'
wb.save(out_path)
print(f'\nSaved: {out_path}')
